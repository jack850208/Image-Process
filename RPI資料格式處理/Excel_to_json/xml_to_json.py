# -*- coding: utf-8 -*-
"""
Created on Thu May 27 17:57:11 2021

@author: AIC
"""

from openpyxl import load_workbook as xml_load
from openpyxl import Workbook as xml_workbook
import json


def workbook_load(wb_name:str):
    wb = xml_load(wb_name)
    return wb
    
def list_to_dict(tag_array:list, value_array:list):
    
    new_dict = {}

    
    for index in range(len(tag_array)):
        new_dict[tag_array[index]] = value_array[index]

    return new_dict

def dict_add(new_dict:dict, add_dict:dict, index_dict):
    new_dict[index_dict] = add_dict
    return new_dict
    
         

if __name__ == '__main__':
    wb = xml_load(r"D:\109 Master Jack\Python\期末專題\Excel_file\Test01.xlsx")
    
    title = []
    data_index = []
    row_temp = [[]]
    cnt = 0
    
    # read excel data title
    for item in wb[wb.sheetnames[1]][1]:
        title.append(item.value)
    # read excel data index
    for item in wb[wb.sheetnames[1]]:
        data_index.append(item[0].value)
    
    # read excel data to list
    for item in wb[wb.sheetnames[1]]:
        for num in item:
            if(len(row_temp[cnt])==len(item)):
                row_temp.append([])
                cnt = cnt + 1
            row_temp[cnt].append(num.value)

    
    # create dictionary to store data
    new_dict = {}
    dict_catch = list_to_dict(title, row_temp[1])
    for index in range(len(row_temp)):
        list_catch = row_temp[index]
        dict_catch = list_to_dict(title, list_catch)
        index_name = "chip" + str(index)
        new_dict = dict_add(new_dict, dict_catch, index_name)
        
        
    json_out = json.dumps(new_dict)
    fp = open(r"excel_to_json.json",r"w")
    fp.write(json_out)
    fp.close()